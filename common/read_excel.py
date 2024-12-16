from openpyxl import load_workbook


class Read_excel:
    case = {"menu1": "一级菜单", "menu2": "二级菜单",
            "purpose": "测试目的",
            "test_items": "测试项", "test_child": "测试子项", "step": "测试步骤",
            "result": "预期结果", "condition": "预置条件", "scene": "组合维度/覆盖场景",
            "priority": "重要程度", "direction": "正反例", "case_type": "用例类型", "remark": "备注"}

    data = {
        "title": "",
        "topics": []
    }

    @classmethod
    def read_excel(cls, excel_file):
        "excel数据提取"
        total_cases = []
        workbook = load_workbook(excel_file)
        worksheet = workbook.worksheets[0]

        list_rows = None
        "获取第一行所有数据的值"
        for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            list_rows = list(row)
        "获取菜单索引"
        menu1_nun = list_rows.index(cls.case["menu1"])
        menu2_num = list_rows.index(cls.case["menu2"])
        purpose_num = list_rows.index(cls.case["purpose"])
        test_items_num = list_rows.index(cls.case["test_items"])
        test_child_num = list_rows.index(cls.case["test_child"])
        step_num = list_rows.index(cls.case["step"])
        result_num = list_rows.index(cls.case["result"])
        condition_num = list_rows.index(cls.case["condition"])
        scene_num = list_rows.index(cls.case["scene"])
        priority_num = list_rows.index(cls.case["priority"])
        direction_num = list_rows.index(cls.case["direction"])
        case_type_num = list_rows.index(cls.case["case_type"])
        remark_num = list_rows.index(cls.case["remark"])

        cls.data["title"] = f"{worksheet['A2'].value}_{worksheet['B2'].value}"
        # 从第二行开始,设置value_only=True只获取单元格的值
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                break

            case = {"menu1": f"{cls.case['menu1']}：{row[menu1_nun]}",
                    "menu2": f"{cls.case['menu2']}：{row[menu2_num]}",
                    "purpose": f"{cls.case['purpose']}：{row[purpose_num]}",
                    "test_items": f"{cls.case['test_items']}：{row[test_items_num]}",
                    "test_child": f"{cls.case['test_child']}：{row[test_child_num]}",
                    "step": f"{cls.case['step']}：{row[step_num]}",
                    "result": f"{cls.case['result']}：{row[result_num]}",
                    "condition": f"{cls.case['condition']}：{row[condition_num]}",
                    "scene": f"覆盖场景：{row[scene_num]}",
                    "priority": f"{cls.case['priority']}：{row[priority_num]}",
                    "direction": f"{cls.case['direction']}：{row[direction_num]}",
                    "case_type": f"{cls.case['case_type']}：{row[case_type_num]}",
                    "remark": f"{cls.case['remark']}：{row[remark_num]}"
                    }

            # case = {"menu1": row[4], "menu2": row[5], "purpose": row[9],
            #         "test_items": row[6], "test_child": row[7], "step": row[14],
            #         "result": row[15], "condition": row[11], "scene": row[12],
            #         "priority": row[10], "direction": row[19], "case_type": row[20], "remark": row[21]}
            total_cases.append(case)
        return total_cases

    def process_data(self, excel_file):
        "对提取的数据进行重新整合，调整数据结构"

        total_cases = Read_excel.read_excel(excel_file)
        # total_cases=self.read_excel(excel_file)

        data = Read_excel.data
        for case in total_cases:
            menu1 = case['menu1']  # 一级菜单
            menu2 = case['menu2']  # 二级菜单
            purpose = case['purpose']  # 测试目的
            test_items = case['test_items']  # 测试项
            test_child = case['test_child']  # 测试子项
            step = case['step']  # 测试步骤
            result = case['result']  # 预期结果
            priority = case["priority"]  # 重要程度
            condition = case["condition"]  # 预置条件
            scene = case["scene"]  # 覆盖场景
            direction = case["direction"]  # 正反例
            case_type = case["case_type"]  # 用例类型
            remark = case["remark"]  # 备注

            # 检查是否存在相同数据
            is_exist = any(
                item['title'] == menu1
                and item['topics'][0]['title'] == menu2
                # and item['topics'][0]['topics'][0]['title'] == purpose
                and item['topics'][0]['topics'][0]['title'] == test_items
                for item in data['topics']
            )

            if is_exist:
                # 存在相同数据，在最后一个 topics 中添加数据
                for item in data['topics']:
                    if (item['title'] == menu1
                            and item['topics'][0]['title'] == menu2
                            # and item['topics'][0]['topics'][0]['title'] == purpose
                            and item['topics'][0]['topics'][0]['title'] == test_items):
                        item['topics'][0]['topics'][0]['topics'].append(
                            {'title': test_child, 'topics':
                                [{'title': step, 'topics':
                                    [{'title': result, 'topics':
                                        [{'priority': priority},
                                         {'condition': condition},
                                         {'scene': scene},
                                         {'direction': direction},
                                         {'case_type': case_type},
                                         {'remark': remark}
                                         ]}]}]})
                        break
            else:
                # 不存在相同数据，判断从哪个节点开始不存在，进行添加
                current_level = data['topics']
                if menu1:
                    if any(item['title'] == menu1 for item in current_level):
                        current_level = next(item['topics'] for item in current_level if item['title'] == menu1)
                    else:
                        current_level.append({"title": menu1, "topics": []})
                        current_level = current_level[-1]['topics']
                if menu2:
                    if any(item['title'] == menu2 for item in current_level):
                        current_level = next(item['topics'] for item in current_level if item['title'] == menu2)
                    else:
                        current_level.append({"title": menu2, "topics": []})
                        current_level = current_level[-1]['topics']
                # if purpose:
                #     if any(item['title'] == purpose for item in current_level):
                #         current_level = next(item['topics'] for item in current_level if item['title'] == purpose)
                #     else:
                #         current_level.append({"title": purpose, "topics": []})
                #         current_level = current_level[-1]['topics']
                if test_items:
                    if any(item['title'] == test_items for item in current_level):
                        current_level = next(item['topics'] for item in current_level if item['title'] == test_items)
                    else:
                        current_level.append({"title": test_items, "topics": []})
                        current_level = current_level[-1]['topics']

                # 添加测试子项、步骤、预期结果
                current_level.append(
                    {'title': test_child, 'topics':
                        [{'title': step, 'topics':
                            [{'title': result, 'topics':
                                [{'priority': priority},
                                 {'condition': condition},
                                 {'scene': scene},
                                 {'direction': direction},
                                 {'case_type': case_type},
                                 {'remark': remark}]
                              }]}]})
        return data


if __name__ == "__main__":
    excel_file = r'D:\PythonProject\ConversionCase\temp_folder\excel_to_xmind-demo.xlsx'
    print("Read_excel.data=", Read_excel.data)
    info = Read_excel()
    # info.read_excel(excel_file)
    # print("Read_excel.data=",Read_excel.data)
    data = info.process_data(excel_file)
    print("data=", data)
