from openpyxl import load_workbook
import os
import shutil
import time
from config import setting
from datetime import datetime


class Read_need_excel:

    case = {"menu1": "一级菜单", "menu2": "二级菜单", "purpose": "测试目的",
            "test_items": "测试项", "test_child": "测试子项", "step": "测试步骤",
            "result": "预期结果", "condition": "预置条件", "scene": "组合维度/覆盖场景",
            "priority": "重要程度", "direction": "正反例", "case_type": "用例类型", "remark": "备注"}

    data = {
        "title": "",
        "topics": []
    }

    def __init__(self, num="None"):
        if str(num) == "3":  #excel->定制xmind格式
            "如果文件夹存在，则删除,清除缓存数据"
            if os.path.exists(setting.temp_folder):
                "文件夹删除"
                shutil.rmtree(setting.temp_folder)
            time.sleep(0.5)
            if not os.path.exists(setting.temp_folder):
                "创建文件夹"
                os.mkdir(setting.temp_folder)
        else:  #标准xmind->excel->定制xmind格式
            if not os.path.exists(setting.target_case_file):
                print("中转文件不存在，请检查标准xmind转excel实现是否成功")

    @classmethod
    def read_excel(cls, excel_file):
        "excel数据提取"
        total_cases = []
        workbook = load_workbook(excel_file)
        worksheet = workbook.worksheets[0]

        list_rows = None
        "获取第一行所有数据的值"
        for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            list_rows = list(row)
        "获取菜单索引"
        menu1_nun = list_rows.index(cls.case["menu1"])
        menu2_num = list_rows.index(cls.case["menu2"])
        test_items_num = list_rows.index(cls.case["test_items"])
        test_child_num = list_rows.index(cls.case["test_child"])
        step_num = list_rows.index(cls.case["step"])
        result_num = list_rows.index(cls.case["result"])

        cls.data["title"] = f"{worksheet['A2'].value}_{worksheet['B2'].value}"
        # 从第二行开始,设置value_only=True只获取单元格的值
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                break

            case = {
                "point1": f"{row[menu1_nun]}",
                "point2":f"{row[menu2_num]}",
                "point3": f"{row[test_items_num]}",
                "point4": f"{row[test_child_num]}",
                "point5": f"{row[step_num]}",
                "point6": f"{row[result_num]}"
            }

            total_cases.append(case)
        return total_cases

    def process_data(self, excel_file=setting.target_case_file):
        "对提取的数据进行重新整合，调整数据结构"

        total_cases = Read_need_excel.read_excel(excel_file)
        # total_cases=self.read_excel(excel_file)
        data = Read_need_excel.data

        for case in total_cases:
            point1 = case["point1"]  #优先级1，自定义项"一级菜单“
            point2 = case["point2"]  #优先级2，自定义项”二级菜单“
            point3 = case["point3"]  #优先级3，测试项
            point4 = case["point4"]  #优先级4，测试子项
            point5 = case["point5"]  #优先级5，测试步骤
            point6 = case["point6"]  #优先级6，预期结果

            # 检查是否存在相同数据
            is_exist = any(
                item['title'] == point1
                and item['topics'][0]['title'] == point2
                and item['topics'][0]['topics'][0]['title'] == point3
                and item['topics'][0]['topics'][0]['topics'][0]['title'] == point4
                for item in data['topics']
            )

            if is_exist:
                # 存在相同数据，在最后一个 topics 中添加数据
                for item in data['topics']:
                    if (item['title'] == point1
                            and item['topics'][0]['title'] == point2
                            and item['topics'][0]['topics'][0]['title'] == point3
                            and item['topics'][0]['topics'][0]['topics'][0]['title'] == point4):
                        item['topics'][0]['topics'][0]['topics'][0]['topics'].append(
                            {'title': point5, 'topics':
                                [{'title': point6}]})
                        break
            else:
                # 不存在相同数据，判断从哪个节点开始不存在，进行添加
                current_level = data['topics']
                if point1:
                    if any(item['title'] == point1 for item in current_level):
                        current_level = next(item['topics'] for item in current_level if item['title'] == point1)
                    else:
                        current_level.append({"title": point1, "topics": []})
                        current_level = current_level[-1]['topics']
                if point2:
                    if any(item['title'] == point2 for item in current_level):
                        current_level = next(item['topics'] for item in current_level if item['title'] == point2)
                    else:
                        current_level.append({"title": point2, "topics": []})
                        current_level = current_level[-1]['topics']
                if point3:
                    if any(item['title'] == point3 for item in current_level):
                        current_level = next(item['topics'] for item in current_level if item['title'] == point3)
                    else:
                        current_level.append({"title": point3, "topics": []})
                        current_level = current_level[-1]['topics']
                if point4:
                    if any(item['title'] == point4 for item in current_level):
                        current_level = next(item['topics'] for item in current_level if item['title'] == point4)
                    else:
                        current_level.append({"title": point4, "topics": []})
                        current_level = current_level[-1]['topics']

                # 添加测试子项、步骤、预期结果
                current_level.append(
                    {'title': point5, 'topics':
                        [{'title': point6}]
                     })
        return data


if __name__ == "__main__":
    excel_file = r'D:\PythonProject\ConversionCase\source_folder\excel_to_xmind-demo.xlsx'
    print("Read_excel.data=", Read_need_excel.data)
    info = Read_need_excel()
    # info.read_excel(excel_file)
    # print("Read_excel.data=",Read_excel.data)
    data = info.process_data(excel_file)
    print("data=", data)
