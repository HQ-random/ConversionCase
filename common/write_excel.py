import os
import time

from config import setting
import shutil
from openpyxl import load_workbook
from datetime import datetime


class Write_excel:

    def __init__(self):
        "如果文件夹存在，则删除,清除缓存数据"
        if os.path.exists(setting.temp_folder):
            "文件夹删除"
            shutil.rmtree(setting.temp_folder)
        time.sleep(0.5)
        if not os.path.exists(setting.temp_folder):
            "创建文件夹"
            os.mkdir(setting.temp_folder)
            "复制源文件至指定目录下，然后在被复制后的文件上进行操作"
            shutil.copy(setting.source_template_file, setting.target_case_file)
            shutil.copy(setting.source_point_file, setting.target_point_file)

    def menu_num(self, total_case):
        "统计一级菜单+二级菜单的数量，确定用例名称升序时的最大数据"
        case_menu_num = {}
        temp_case = set()
        new_case = set()  # 将 new_case 的初始化移到循环体外部
        for x in total_case:
            case_num = f"{x['menu1']}_{x['menu2']}_{x['test_items']}"
            new_case.add(case_num)
            case_menu = f"{x['menu1']}_{x['menu2']}"
            if new_case.issubset(temp_case):
                "判断new_case是否为temp_case的子集，间接判断一级菜单+二级菜单下的用例数量"
                case_menu_num[case_menu] = case_menu_num[case_menu] + 1
            else:
                temp_case.add(case_num)
                if case_menu_num.get(case_menu, -1) != -1:
                    case_menu_num[case_menu] = case_menu_num[case_menu] + 1
                else:
                    case_menu_num[case_menu] = 1
            # time.sleep(0.1)
        # print("case_menu_num=", case_menu_num, "\ntemp_case=", temp_case)
        return case_menu_num

    def write_excel_case(self, total_cases):
        "测试用例模板文件写入"
        "每个二级菜单下的用例数"
        case_menu_num = self.menu_num(total_cases)
        workbook = load_workbook(setting.target_case_file)
        "获取第一个sheet页"
        worksheet = workbook.worksheets[0]

        "获取第二个sheet页"
        second_worksheet = workbook.worksheets[1]
        dtime = datetime.now().strftime("%Y%m%d")
        "临时写入用例序号统计"
        case_menu_count = {}

        for x in range(len(total_cases)):
            "系统名称"
            worksheet[f"A{x + 2}"] = second_worksheet["A2"].value
            "版本号"
            worksheet[f"B{x + 2}"] = second_worksheet["B2"].value
            "需求摘要"
            worksheet[f"C{x + 2}"] = second_worksheet["C2"].value
            "交易日"
            worksheet[f"D{x + 2}"] = total_cases[x]["trading_day"]
            "一级菜单"
            worksheet[f"E{x + 2}"] = total_cases[x]["menu1"]
            "二级菜单"
            worksheet[f"F{x + 2}"] = total_cases[x]["menu2"]
            "测试项"
            worksheet[f"G{x + 2}"] = total_cases[x]["test_items"]
            "测试子项"
            worksheet[f"H{x + 2}"] = total_cases[x]["test_child"]

            "临时菜单"
            case_menu = f"{total_cases[x]['menu1']}_{total_cases[x]['menu2']}"
            if case_menu_num.get(case_menu, -1) != -1:
                if case_menu_count.get(case_menu,-999) != -999:
                    case_menu_count[case_menu] = case_menu_count[case_menu] + 1
                else:
                    case_menu_count[case_menu] = 1
                "用例名称"
                worksheet[f"I{x + 2}"] = f"{total_cases[x]['menu1']}_{total_cases[x]['menu2']}_{case_menu_count[case_menu]}"
                # print(f"{total_cases[x]['menu1']}_{total_cases[x]['menu2']}_{case_menu_count[case_menu]}")
                case_menu_num[case_menu] = case_menu_num[case_menu] - 1
                if case_menu_num[case_menu] == -1:
                    print(case_menu_num, "：一级菜单+二级菜单数量统计错误，请重新检查函数menu_num()")
            else:
                print("未知的菜单：", case_menu)

            "测试目的"
            worksheet[f"J{x + 2}"] = total_cases[x]["purpose"]
            "重要程度"
            worksheet[f"K{x + 2}"] = total_cases[x]["priority"]
            "预置条件"
            worksheet[f"L{x + 2}"] = total_cases[x]["condition"]
            "覆盖场景"
            worksheet[f"M{x + 2}"] = total_cases[x]["scene"]
            "测试用例数"
            worksheet[f"N{x + 2}"] = total_cases[x]["case_num"]
            "测试步骤"
            worksheet[f"O{x + 2}"] = total_cases[x]["step"]
            "预期结果"
            worksheet[f"P{x + 2}"] = total_cases[x]["result"]
            "编写日期,默认当前系统时间"
            worksheet[f"Q{x + 2}"] = dtime
            "编写人"
            worksheet[f"R{x + 2}"] = total_cases[x]["create_name"]
            "是否基线用例"
            worksheet[f"S{x + 2}"] = total_cases[x]["baseline"]
            "正反例"
            worksheet[f"T{x + 2}"] = total_cases[x]["direction"]
            "用例类型"
            worksheet[f"U{x + 2}"] = total_cases[x]["case_type"]
            "备注"
            worksheet[f"V{x + 2}"] = total_cases[x]["remark"]
        "保存"
        workbook.save(setting.target_case_file)

    def write_excel_point(self, total_cases):
        workbook = load_workbook(setting.target_point_file)
        "获取第一个sheet页"
        worksheet = workbook.worksheets[0]
        "获取第二个sheet页"
        second_worksheet = workbook.worksheets[1]
        case_points = set()
        for x in range(len(total_cases)):
            "直接使用测试目的，并去重"
            case_point = f'{total_cases[x]["purpose"]}'
            case_points.add(case_point)
        # print("len(case_points):",len(case_points),"case_points:",case_points)
        n = 0
        for x in case_points:
            "系统名称"
            worksheet[f"A{n + 2}"] = second_worksheet["A2"].value
            "二级菜单"
            worksheet[f"B{n + 2}"] = x.split("_")[0][2:]
            "功能点"
            worksheet[f"C{n + 2}"] = x
            n += 1
        time.sleep(1)
        "保存"
        workbook.save(setting.target_point_file)


if __name__ == "__main__":
    w = Write_excel()
